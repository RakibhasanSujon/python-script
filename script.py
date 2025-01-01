from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from datetime import datetime

def get_day_of_week():
    today = datetime.today()
    return today.strftime('%A')

def get_keywords(day):

    if day == "Sunday":
        return ["Dhaka","Sunday","Baby","School","Cricket","Money","Int","Look","Hello","By"]
    elif day == "Monday":
        return  ["Dhaka","Dates","Baby","School","Cricket","Money","Int","Look","Hello","By"]
    elif day == "Tuesday":
        return  ["Dhaka","Saturday","Baby","School","Cricket","Money","Int","Hubby","Hello","By"]
    elif day == "Wednesday":
        return  ["Dhaka","Saturday","Baby","School","Cricket","Pathao","Int","Look","Hello","By"]
    elif day == "Thursday":
        return  ["Dhaka","Saturday","Baby","School","Cricket","Money","Int","Goods","Hello","By"]
    elif day == "Friday":
        return  ["Dhaka","Saturday","Baby","School","Geography","Money","Int","Look","Hello","By"]
    elif day == "Saturday":
        return  ["Dhaka","Saturday","Baby","School","Cricket","Money","Int","Look","Hello","By"]

def google_search(keyword):

    driver = webdriver.Chrome()
    driver.get(f"https://www.google.com/search?q={keyword}")

    try:
        # Wait for search results to load (adjust as needed)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".g"))
        )

        # Extract search results (adjust selectors based on Google's HTML structure)
        results = driver.find_elements(By.CSS_SELECTOR, ".g")

        longest_option = None
        shortest_option = None
        max_length = 0
        min_length = float('inf')

        for result in results:
            text = result.text
            if text:
                length = len(text)
                if length > max_length:
                    max_length = length
                    longest_option = text
                if length < min_length:
                    min_length = length
                    shortest_option = text

    except Exception as e:
        print(f"Error searching for {keyword}: {e}")
        longest_option = None
        shortest_option = None

    finally:
        driver.quit()
        return longest_option, shortest_option

def save_to_excel(data):

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Set column headers
    sheet.cell(row=1, column=1).value = "Keyword"
    sheet.cell(row=1, column=2).value = "Longest Option"
    sheet.cell(row=1, column=3).value = "Shortest Option"

    # Write data to the sheet
    for i, row in enumerate(data):
        sheet.cell(row=i+2, column=1).value = row[0]  # Keyword
        sheet.cell(row=i+2, column=2).value = row[1]  # Longest Option
        sheet.cell(row=i+2, column=3).value = row[2]  # Shortest Option

    workbook.save("google_search_results.xlsx")

def main():
    day = get_day_of_week()
    keywords = get_keywords(day)

    results = []
    for keyword in keywords:
        longest, shortest = google_search(keyword)
        results.append((keyword, longest, shortest))

    save_to_excel(results)

if __name__ == "__main__":
    main()